"""Parse AISC Shapes Database v16.0 into a production-grade JSON knowledge base.

Reads the full AISC v16.0 shapes database Excel file and exports a structured
JSON file with all section families and key engineering properties.
"""

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

# Paths
REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = REPO_ROOT / "Eagle Data" / "BOT TRAINING" / "Engineering" / "aisc-shapes-database-v16.0.xlsx"
OUTPUT_FILE = REPO_ROOT / "data" / "standards" / "aisc_shapes.json"

# Column indices in the "Database v16.0" sheet
COL = {
    "Type": 0,
    "EDI_Std_Nomenclature": 1,
    "AISC_Manual_Label": 2,
    "T_F": 3,
    "W": 4,
    "A": 5,
    "d": 6,
    "ddet": 7,
    "Ht": 8,
    "h": 9,
    "OD": 10,
    "bf": 11,
    "bfdet": 12,
    "B": 13,
    "b": 14,
    "ID": 15,
    "tw": 16,
    "twdet": 17,
    "tf": 19,
    "tfdet": 20,
    "t": 21,
    "tnom": 22,
    "tdes": 23,
    "kdes": 24,
    "kdet": 25,
    "k1": 26,
    "x": 27,
    "y": 28,
    "Ix": 38,
    "Zx": 39,
    "Sx": 40,
    "rx": 41,
    "Iy": 42,
    "Zy": 43,
    "Sy": 44,
    "ry": 45,
    "Iz": 46,
    "rz": 47,
    "J": 49,
    "Cw": 50,
    "C": 51,
    "rts": 74,
    "ho": 75,
}

# Dash character used for N/A values in the spreadsheet
DASH = "\u2013"  # en-dash


def _val(row, col_name):
    """Extract a numeric value from a row, returning None for dashes/blanks."""
    v = row[COL[col_name]]
    if v is None or v == DASH or v == "–" or v == "-":
        return None
    try:
        return round(float(v), 6)
    except (ValueError, TypeError):
        return None


def _fy_for_shape(shape_type, t_f_flag):
    """Determine Fy (ksi) based on shape type and T_F flag.

    Per AISC 16th Ed:
    - W/M/S/HP: Fy=50 (A992), T_F='T' means column per AISC compactness = Fy=50
    - C/MC: Fy=36 (A36)
    - L: Fy=36 (A36)
    - HSS (round/square/rect): Fy=46 (A500 Gr. C)
    - PIPE: Fy=35 (A53 Gr. B) for STD/XS, Fy=35 for XXS
    - WT/MT/ST: same as parent (50/36/36)
    """
    if shape_type in ("W", "M", "S", "HP"):
        return 50
    if shape_type in ("WT", "MT"):
        return 50
    if shape_type == "ST":
        return 36
    if shape_type in ("C", "MC"):
        return 36
    if shape_type == "L":
        return 36
    if shape_type == "HSS":
        return 46
    if shape_type == "PIPE":
        return 35
    if shape_type == "2L":
        return 36
    return None


def _classify_hss(label):
    """Classify HSS as square, rectangular, or round based on label."""
    if not label:
        return "HSS_rect"
    label_upper = label.upper().replace(" ", "")
    # Round HSS: HSS followed by a single dimension and wall thickness
    # e.g., HSS20.000X0.500 (OD x t)
    # Rectangular/Square: HSS HxBxt where H and B are present
    # If Ht == B, it's square
    parts = label_upper.replace("HSS", "").split("X")
    if len(parts) == 2:
        # Could be round: OD x t
        return "HSS_round"
    if len(parts) >= 3:
        try:
            h = float(parts[0].replace("/", ""))
        except ValueError:
            h = 0
        try:
            b_val = float(parts[1].replace("/", ""))
        except ValueError:
            b_val = 0
        if h == b_val:
            return "HSS_square"
        return "HSS_rect"
    return "HSS_rect"


def _build_w_shape(row):
    """Build a W/M/S/HP shape entry."""
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "d_in": _val(row, "d"),
        "bf_in": _val(row, "bf"),
        "tf_in": _val(row, "tf"),
        "tw_in": _val(row, "tw"),
        "kdes_in": _val(row, "kdes"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "Iy_in4": _val(row, "Iy"),
        "Zy_in3": _val(row, "Zy"),
        "Sy_in3": _val(row, "Sy"),
        "ry_in": _val(row, "ry"),
        "J_in4": _val(row, "J"),
        "Cw_in6": _val(row, "Cw"),
        "rts_in": _val(row, "rts"),
        "ho_in": _val(row, "ho"),
        "Fy_ksi": _fy_for_shape(row[COL["Type"]], row[COL["T_F"]]),
    }


def _build_hss_shape(row):
    """Build an HSS shape entry."""
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "Ht_in": _val(row, "Ht"),
        "B_in": _val(row, "B"),
        "tnom_in": _val(row, "tnom"),
        "tdes_in": _val(row, "tdes"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "Iy_in4": _val(row, "Iy"),
        "Zy_in3": _val(row, "Zy"),
        "Sy_in3": _val(row, "Sy"),
        "ry_in": _val(row, "ry"),
        "J_in4": _val(row, "J"),
        "C_in3": _val(row, "C"),
        "Fy_ksi": 46,
    }


def _build_hss_round_shape(row):
    """Build a round HSS shape entry."""
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "OD_in": _val(row, "OD"),
        "tnom_in": _val(row, "tnom"),
        "tdes_in": _val(row, "tdes"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "J_in4": _val(row, "J"),
        "C_in3": _val(row, "C"),
        "Fy_ksi": 46,
    }


def _build_pipe_shape(row):
    """Build a pipe shape entry."""
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "OD_in": _val(row, "OD"),
        "ID_in": _val(row, "ID"),
        "tnom_in": _val(row, "tnom"),
        "tdes_in": _val(row, "tdes"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "J_in4": _val(row, "J"),
        "C_in3": _val(row, "C"),
        "Fy_ksi": 35,
    }


def _build_channel_shape(row):
    """Build a C or MC channel shape entry."""
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "d_in": _val(row, "d"),
        "bf_in": _val(row, "bf"),
        "tf_in": _val(row, "tf"),
        "tw_in": _val(row, "tw"),
        "kdes_in": _val(row, "kdes"),
        "x_in": _val(row, "x"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "Iy_in4": _val(row, "Iy"),
        "Zy_in3": _val(row, "Zy"),
        "Sy_in3": _val(row, "Sy"),
        "ry_in": _val(row, "ry"),
        "J_in4": _val(row, "J"),
        "Cw_in6": _val(row, "Cw"),
        "Fy_ksi": 36,
    }


def _build_angle_shape(row):
    """Build an L angle shape entry."""
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "d_in": _val(row, "d"),
        "b_in": _val(row, "b"),
        "t_in": _val(row, "t"),
        "kdes_in": _val(row, "kdes"),
        "x_in": _val(row, "x"),
        "y_in": _val(row, "y"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "Iy_in4": _val(row, "Iy"),
        "Zy_in3": _val(row, "Zy"),
        "Sy_in3": _val(row, "Sy"),
        "ry_in": _val(row, "ry"),
        "Iz_in4": _val(row, "Iz"),
        "rz_in": _val(row, "rz"),
        "J_in4": _val(row, "J"),
        "Fy_ksi": 36,
    }


def _build_tee_shape(row):
    """Build a WT/MT/ST tee shape entry."""
    shape_type = row[COL["Type"]]
    return {
        "designation": row[COL["AISC_Manual_Label"]],
        "weight_plf": _val(row, "W"),
        "area_in2": _val(row, "A"),
        "d_in": _val(row, "d"),
        "bf_in": _val(row, "bf"),
        "tf_in": _val(row, "tf"),
        "tw_in": _val(row, "tw"),
        "Ix_in4": _val(row, "Ix"),
        "Zx_in3": _val(row, "Zx"),
        "Sx_in3": _val(row, "Sx"),
        "rx_in": _val(row, "rx"),
        "Iy_in4": _val(row, "Iy"),
        "Zy_in3": _val(row, "Zy"),
        "Sy_in3": _val(row, "Sy"),
        "ry_in": _val(row, "ry"),
        "J_in4": _val(row, "J"),
        "Fy_ksi": _fy_for_shape(shape_type, row[COL["T_F"]]),
    }


def parse_aisc_database():
    """Parse the full AISC v16.0 database and return structured data."""
    print(f"Reading: {EXCEL_FILE}")
    wb = load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)
    ws = wb["Database v16.0"]

    families = {
        "W": [],
        "M": [],
        "S": [],
        "HP": [],
        "C": [],
        "MC": [],
        "L": [],
        "WT": [],
        "MT": [],
        "ST": [],
        "HSS_square": [],
        "HSS_rect": [],
        "HSS_round": [],
        "pipe": [],
        "2L": [],
    }

    skipped = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header

        shape_type = row[COL["Type"]]
        if shape_type is None:
            skipped += 1
            continue

        label = row[COL["AISC_Manual_Label"]]

        if shape_type in ("W", "M", "S", "HP"):
            families[shape_type].append(_build_w_shape(row))
        elif shape_type == "HSS":
            hss_class = _classify_hss(label)
            if hss_class == "HSS_round":
                families["HSS_round"].append(_build_hss_round_shape(row))
            else:
                families[hss_class].append(_build_hss_shape(row))
        elif shape_type == "PIPE":
            families["pipe"].append(_build_pipe_shape(row))
        elif shape_type in ("C", "MC"):
            families[shape_type].append(_build_channel_shape(row))
        elif shape_type == "L":
            families["L"].append(_build_angle_shape(row))
        elif shape_type in ("WT", "MT", "ST"):
            families[shape_type].append(_build_tee_shape(row))
        elif shape_type == "2L":
            families["2L"].append(_build_angle_shape(row))
        else:
            skipped += 1
            print(f"  WARNING: Unknown shape type '{shape_type}' at row {i + 1}")

    wb.close()

    total = sum(len(v) for v in families.values())

    result = {
        "metadata": {
            "source": "AISC Shapes Database v16.0",
            "reference": "AISC Steel Construction Manual, 16th Edition",
            "parsed_date": str(date.today()),
            "total_sections": total,
            "skipped_rows": skipped,
            "notes": {
                "Fy_values": {
                    "W_M_S_HP": "50 ksi (ASTM A992)",
                    "C_MC": "36 ksi (ASTM A36)",
                    "L": "36 ksi (ASTM A36)",
                    "HSS": "46 ksi (ASTM A500 Gr. C)",
                    "PIPE": "35 ksi (ASTM A53 Gr. B)",
                },
                "units": {
                    "weight_plf": "lb/ft",
                    "area_in2": "in^2",
                    "d_in": "in",
                    "I_in4": "in^4",
                    "S_in3": "in^3",
                    "Z_in3": "in^3",
                    "r_in": "in",
                    "J_in4": "in^4",
                    "Cw_in6": "in^6",
                    "Fy_ksi": "ksi",
                },
            },
        },
        "families": families,
    }

    return result


def run_spot_checks(data):
    """Run spot checks against known AISC values."""
    print("\n=== SPOT CHECKS ===")
    checks = []

    # W14x22: d=13.7, Ix=199, Sx=29.0, Fy=50
    w14x22 = None
    for s in data["families"]["W"]:
        if s["designation"] == "W14X22":
            w14x22 = s
            break
    if w14x22:
        checks.append(("W14X22 d", w14x22["d_in"], 13.7))
        checks.append(("W14X22 Ix", w14x22["Ix_in4"], 199.0))
        checks.append(("W14X22 Sx", w14x22["Sx_in3"], 29.0))
        checks.append(("W14X22 Fy", w14x22["Fy_ksi"], 50))
    else:
        print("  FAIL: W14X22 not found!")

    # W8x31: d=8.00, Ix=110, Sx=27.5
    w8x31 = None
    for s in data["families"]["W"]:
        if s["designation"] == "W8X31":
            w8x31 = s
            break
    if w8x31:
        checks.append(("W8X31 d", w8x31["d_in"], 8.00))
        checks.append(("W8X31 Ix", w8x31["Ix_in4"], 110.0))
        checks.append(("W8X31 Sx", w8x31["Sx_in3"], 27.5))
    else:
        print("  FAIL: W8X31 not found!")

    # HSS6x6x1/4: A=5.24, Ix from database (v16.0 value)
    hss6 = None
    for family_key in ("HSS_square", "HSS_rect"):
        for s in data["families"][family_key]:
            if s["designation"] == "HSS6X6X1/4":
                hss6 = s
                break
    if hss6:
        checks.append(("HSS6X6X1/4 A", hss6["area_in2"], 5.24))
        checks.append(("HSS6X6X1/4 Ix", hss6["Ix_in4"], 28.6))
        print(
            f"  NOTE: HSS6X6X1/4 Ix=28.6 per v16.0 database "
            f"(task spec cited Ix=40.1 which does not match v16.0)"
        )
    else:
        print("  FAIL: HSS6X6X1/4 not found!")

    # Pipe 4 STD: OD=4.500, Ix from database
    pipe4 = None
    for s in data["families"]["pipe"]:
        if s["designation"] == "Pipe4STD":
            pipe4 = s
            break
    if pipe4:
        checks.append(("Pipe4STD OD", pipe4["OD_in"], 4.5))
        checks.append(("Pipe4STD Ix", pipe4["Ix_in4"], 6.82))
        print(
            f"  NOTE: Pipe4STD Ix=6.82 per v16.0 database "
            f"(task spec cited Ix=7.23 which does not match v16.0)"
        )
    else:
        print("  FAIL: Pipe4STD not found!")

    all_pass = True
    for name, actual, expected in checks:
        if actual is None:
            print(f"  FAIL: {name} = None (expected {expected})")
            all_pass = False
        elif abs(float(actual) - float(expected)) > 0.01:
            print(f"  FAIL: {name} = {actual} (expected {expected})")
            all_pass = False
        else:
            print(f"  PASS: {name} = {actual}")

    return all_pass


def main():
    data = parse_aisc_database()

    # Write JSON
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"\nWritten: {OUTPUT_FILE}")
    print(f"File size: {OUTPUT_FILE.stat().st_size / 1024:.1f} KB")

    # Report counts
    print("\n=== FAMILY COUNTS ===")
    for family, shapes in data["families"].items():
        if shapes:
            print(f"  {family}: {len(shapes)}")
    print(f"  TOTAL: {data['metadata']['total_sections']}")

    # Spot checks
    all_pass = run_spot_checks(data)

    if all_pass:
        print("\nAll spot checks PASSED.")
    else:
        print("\nSome spot checks FAILED. Review above.")
        sys.exit(1)


if __name__ == "__main__":
    main()
