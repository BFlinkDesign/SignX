from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
from datetime import datetime
import json
import os
from pathlib import Path
import logging
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import PyPDF2
from io import BytesIO
import zipfile
from openpyxl.cell.cell import MergedCell

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Data models
class WorkOrder(BaseModel):
    work_order: str
    date: str
    part_number: str
    description: str
    quantity: Optional[int] = None
    labor_hours: Optional[float] = None
    department: Optional[str] = None

class ProcessingResult(BaseModel):
    success: bool
    data: Optional[List[WorkOrder]] = None
    errors: Optional[List[str]] = None
    stats: Optional[Dict[str, Any]] = None

# Enhanced pattern recognition
class PatternRecognizer:
    def __init__(self):
        self.work_order_patterns = [
            r'^[A-Z0-9-]+$',
            r'^WO[A-Z0-9-]+$',
            r'^[A-Z]{2,3}\d{4,6}$',
            r'^\d{4,6}[A-Z]{1,2}$',
            r'Work\s*Order\s*[#:]?\s*(\d{4,6})',
            r'WO[-\s]?(\d{4,6})',
            r'(\d{4,6})[-\s]?[A-Z]?$',
            r'[A-Z]{2,3}-(\d{4,6})',
            r'(\d{4,6})-[A-Z]{2,3}$',
            r'[A-Z]{2,3}\d{4,6}[A-Z]{0,2}$'
        ]
        self.date_patterns = [
            r'^\d{1,2}/\d{1,2}/\d{2,4}$',
            r'^\d{4}-\d{2}-\d{2}$',
            r'^\d{1,2}-\d{1,2}-\d{2,4}$',
            r'^[A-Za-z]{3}\s\d{1,2},\s\d{4}$',
            r'^\d{1,2}\s+[A-Za-z]{3,}\s+\d{4}$',
            r'^\d{1,2}\s+\d{1,2}\s+\d{4}$',
            r'^\d{4}/\d{2}/\d{2}$',
            r'^\d{2}/\d{2}/\d{4}$'
        ]
        self.part_number_patterns = [
            r'^[A-Z0-9-]+$',
            r'^[A-Z]{2,3}\d{4,6}[A-Z]{0,2}$',
            r'^P[A-Z0-9-]+$',
            r'^[A-Z0-9-]+\.[A-Z0-9]{1,3}$',
            r'Part\s*Number\s*[:]?\s*([A-Z0-9-_\.]+)',
            r'P/N\s*[:]?\s*([A-Z0-9-_\.]+)',
            r'[A-Z]{2,3}-[A-Z0-9-]+$',
            r'[A-Z0-9-]+-[A-Z]{2,3}$',
            r'[A-Z]{2,3}\d{4,6}[A-Z]{0,2}\.\d{1,3}$'
        ]
        self.compiled_patterns = {
            'work_order': [re.compile(p) for p in self.work_order_patterns],
            'date': [re.compile(p) for p in self.date_patterns],
            'part_number': [re.compile(p) for p in self.part_number_patterns]
        }

    def extract_work_order(self, text: str) -> Optional[str]:
        for pattern in self.compiled_patterns['work_order']:
            match = pattern.search(text)
            if match:
                return match.group(1) if match.groups() else match.group(0)
        return None

    def extract_date(self, text: str) -> Optional[str]:
        for pattern in self.compiled_patterns['date']:
            if pattern.match(text):
                return text
        return None

    def extract_part_number(self, text: str) -> Optional[str]:
        for pattern in self.compiled_patterns['part_number']:
            match = pattern.search(text)
            if match:
                return match.group(1) if match.groups() else match.group(0)
        return None

# Enhanced data processing
class DataProcessor:
    def __init__(self):
        self.pattern_recognizer = PatternRecognizer()
        self.vectorizer = TfidfVectorizer()
        self.department_codes = {'ES', 'SC', 'PR', 'MFG', 'ENG', 'QA'}

    def process_text(self, text: str) -> ProcessingResult:
        try:
            lines = text.split('\n')
            work_orders = []
            current_part = {}
            errors = []
            stats = {
                'total_lines': len(lines),
                'processed_orders': 0,
                'errors': 0,
                'departments': set()
            }

            for i, line in enumerate(lines, 1):
                try:
                    line = line.strip()
                    if not line:
                        continue

                    # Extract part number
                    part_number = self.pattern_recognizer.extract_part_number(line)
                    if part_number:
                        current_part['part_number'] = part_number
                        continue

                    # Extract work order
                    work_order = self.pattern_recognizer.extract_work_order(line)
                    if work_order:
                        # Extract date from next line if available
                        date = None
                        if i + 1 < len(lines):
                            date = self.pattern_recognizer.extract_date(lines[i + 1].strip())

                        if date:
                            # Extract department code
                            department = next((code for code in self.department_codes 
                                            if code in work_order), None)
                            
                            work_orders.append(WorkOrder(
                                work_order=work_order,
                                date=date,
                                part_number=current_part.get('part_number', ''),
                                description=current_part.get('description', ''),
                                department=department
                            ))
                            stats['processed_orders'] += 1
                            if department:
                                stats['departments'].add(department)

                except Exception as e:
                    errors.append(f"Error processing line {i}: {str(e)}")
                    stats['errors'] += 1

            return ProcessingResult(
                success=len(errors) == 0,
                data=work_orders,
                errors=errors if errors else None,
                stats=stats
            )

        except Exception as e:
            logger.error(f"Error processing data: {str(e)}")
            return ProcessingResult(
                success=False,
                errors=[str(e)]
            )

    def export_to_excel(self, work_orders: List[WorkOrder]) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Work Orders"

        # Add headers
        headers = ['Work Order', 'Date', 'Part Number', 'Description', 'Department']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            if cell is not None and not isinstance(cell, MergedCell):
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

        # Add data
        for row, order in enumerate(work_orders, 2):
            c1 = ws.cell(row=row, column=1)
            if c1 is not None and not isinstance(c1, MergedCell):
                c1.value = order.work_order
            c2 = ws.cell(row=row, column=2)
            if c2 is not None and not isinstance(c2, MergedCell):
                c2.value = order.date
            c3 = ws.cell(row=row, column=3)
            if c3 is not None and not isinstance(c3, MergedCell):
                c3.value = order.part_number
            c4 = ws.cell(row=row, column=4)
            if c4 is not None and not isinstance(c4, MergedCell):
                c4.value = order.description
            c5 = ws.cell(row=row, column=5)
            if c5 is not None and not isinstance(c5, MergedCell):
                c5.value = order.department

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                if cell is not None and not isinstance(cell, MergedCell):
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
            if column and column[0] is not None and not isinstance(column[0], MergedCell) and hasattr(column[0], 'column_letter') and column[0].column_letter is not None:
                ws.column_dimensions[column[0].column_letter].width = max_length + 2

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_to_pdf(self, work_orders: List[WorkOrder]) -> BytesIO:
        output = BytesIO()
        pdf = PyPDF2.PdfWriter()
        
        # Create a simple PDF with the data
        # Note: This is a basic implementation. For better PDF generation,
        # consider using reportlab or weasyprint
        for order in work_orders:
            # Add order details to PDF
            pass

        pdf.write(output)
        output.seek(0)
        return output

# Initialize processor
processor = DataProcessor()

@app.route('/api/process', methods=['POST'])
def process_data():
    try:
        data = request.get_json()
        if not data or 'text' not in data:
            return jsonify({'error': 'No text data provided'}), 400

        result = processor.process_text(data['text'])
        return jsonify(result.dict())

    except Exception as e:
        logger.error(f"Error in process_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def export_data():
    try:
        data = request.get_json()
        if not data or 'work_orders' not in data:
            return jsonify({'error': 'No work orders provided'}), 400

        format_type = data.get('format', 'excel')
        work_orders = [WorkOrder(**wo) for wo in data['work_orders']]

        if format_type == 'excel':
            output = processor.export_to_excel(work_orders)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='work_orders.xlsx'
            )
        elif format_type == 'pdf':
            output = processor.export_to_pdf(work_orders)
            return send_file(
                output,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='work_orders.pdf'
            )
        else:
            return jsonify({'error': 'Unsupported format'}), 400

    except Exception as e:
        logger.error(f"Error in export_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000) 